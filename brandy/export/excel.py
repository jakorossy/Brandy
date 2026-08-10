"""
Multi-sheet Excel workbook builder.

Preserves the Excel export capability from the original codebase
(financial_analyzer.py L374-403 and export.py) but reorganized into
a reusable builder with proper column sizing and number formatting.
"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Header styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# Columns that hold dollar amounts (stored in raw units from XBRL — typically dollars)
_DOLLAR_COLS = {
    "revenue", "cost_of_revenue", "gross_profit", "operating_expenses",
    "operating_income", "net_income", "total_assets", "total_liabilities",
    "shareholders_equity", "cash_and_equivalents", "operating_cash_flow",
    "capital_expenditures", "depreciation_amortization", "tax_expense",
    "pretax_income", "sga_expense",
    "short_term_debt", "long_term_debt", "total_debt", "interest_expense",
    "operating_lease_liability_current", "operating_lease_liability_noncurrent",
    "market_cap",
    "ebitda", "free_cash_flow",
}

# Columns that are ratios/margins — display as percentages
_PERCENT_COLS = {
    "gross_margin", "operating_margin", "net_margin", "roe",
    # WACC sheet
    "risk_free_rate", "equity_risk_premium", "cost_of_equity",
    "pre_tax_cost_of_debt", "tax_rate", "after_tax_cost_of_debt",
    "weight_equity", "weight_debt", "wacc",
}

# Columns that are plain ratios (not percentages)
_RATIO_COLS = {
    "debt_to_equity", "eps_diluted", "beta",
}

# Preferred column order for the Financial Data sheet
_FINANCIAL_COL_ORDER = [
    "ticker", "fiscal_year", "filing_date",
    "revenue", "cost_of_revenue", "gross_profit",
    "gross_profit_source", "gross_profit_notes",
    "operating_expenses", "sga_expense",
    "operating_income", "operating_income_source",
    "depreciation_amortization", "tax_expense", "pretax_income",
    "ebitda", "ebitda_source", "ebitda_notes",
    "net_income", "eps_diluted",
    "total_assets", "total_liabilities", "shareholders_equity",
    "short_term_debt", "long_term_debt",
    "operating_lease_liability_current", "operating_lease_liability_noncurrent",
    "total_debt", "interest_expense",
    "cash_and_equivalents", "operating_cash_flow",
    "capital_expenditures", "free_cash_flow",
    "gross_margin", "operating_margin", "net_margin",
    "roe", "debt_to_equity",
    "data_source",
]

# Preferred column order for the WACC sheet — inputs, then derived outputs.
_WACC_COL_ORDER = [
    "ticker", "fiscal_year",
    "risk_free_rate", "equity_risk_premium", "beta", "cost_of_equity",
    "pre_tax_cost_of_debt", "tax_rate", "after_tax_cost_of_debt",
    "market_cap", "total_debt", "weight_equity", "weight_debt",
    "wacc", "assumptions_note",
]

# openpyxl number format strings
_FMT_DOLLAR = '#,##0'           # 1,234,567  (whole dollars — XBRL values are in dollars)
_FMT_PERCENT = '0.0%'           # 12.3%
_FMT_RATIO = '0.00'             # 1.45


def _reorder_columns(df: pd.DataFrame, preferred_order: list[str]) -> pd.DataFrame:
    """Reorder df columns to match preferred_order, keeping any extra columns at the end."""
    ordered = [c for c in preferred_order if c in df.columns]
    extras = [c for c in df.columns if c not in ordered]
    return df[ordered + extras]


def _apply_number_formats(ws, df: pd.DataFrame):
    """Apply number formats to data cells (not header) based on column type."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in _DOLLAR_COLS:
            fmt = _FMT_DOLLAR
        elif col_name in _PERCENT_COLS:
            fmt = _FMT_PERCENT
        elif col_name in _RATIO_COLS:
            fmt = _FMT_RATIO
        else:
            continue
        for row_idx in range(2, len(df) + 2):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt


def _auto_size_columns(ws, df: pd.DataFrame, wrap_columns: list[str] = None, wrap_width: int = 60):
    """
    Auto-size columns based on content, with optional text wrapping
    for wide columns like commentary.

    Fixes the column-letter bug from the original code (breaks past col Z).
    """
    wrap_columns = wrap_columns or []

    for i, col_name in enumerate(df.columns):
        col_letter = get_column_letter(i + 1)

        if col_name in wrap_columns:
            ws.column_dimensions[col_letter].width = wrap_width
            for row in range(2, len(df) + 2):
                cell = ws[f"{col_letter}{row}"]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            max_width = max(
                df[col_name].astype(str).apply(len).max() if len(df) > 0 else 0,
                len(str(col_name)),
            ) + 2
            ws.column_dimensions[col_letter].width = min(max_width, 40)

    # Style header row
    for i, col_name in enumerate(df.columns):
        cell = ws.cell(row=1, column=i + 1)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, wrap_columns: list[str] = None):
    """Write a DataFrame to a named sheet with formatting."""
    if df.empty:
        logger.info("Skipping empty sheet: %s", sheet_name)
        return

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _apply_number_formats(ws, df)
    _auto_size_columns(ws, df, wrap_columns)

    # Set row height for wrapped content
    if wrap_columns:
        for row_num in range(2, len(df) + 2):
            ws.row_dimensions[row_num].height = 80

    logger.info("Wrote sheet '%s' with %d rows", sheet_name, len(df))


def build_financial_workbook(
    metrics_df: pd.DataFrame,
    commentary_df: pd.DataFrame = None,
    wacc_df: pd.DataFrame = None,
    output_dir: str = "output",
    filename_prefix: str = "financial_analysis",
) -> str:
    """
    Build a financial analysis Excel workbook.

    Sheets:
    - Financial Data: core metrics and ratios (columns reordered, numbers formatted)
    - WACC: assumptions and calculations (if provided)
    - Commentary: AI-generated summaries (if provided)
    """
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"{filename_prefix}_{ts}.xlsx")

    # Reorder columns before writing
    metrics_df = _reorder_columns(metrics_df, _FINANCIAL_COL_ORDER)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        write_sheet(writer, metrics_df, "Financial Data")

        if wacc_df is not None and not wacc_df.empty:
            wacc_df = _reorder_columns(wacc_df, _WACC_COL_ORDER)
            write_sheet(writer, wacc_df, "WACC", wrap_columns=["assumptions_note"])

        if commentary_df is not None and not commentary_df.empty:
            write_sheet(writer, commentary_df, "Commentary", wrap_columns=["commentary"])

    logger.info("Financial workbook saved: %s", path)
    return path


def build_social_workbook(
    posts_df: pd.DataFrame,
    sentiment_df: pd.DataFrame = None,
    accounts_df: pd.DataFrame = None,
    commentary_df: pd.DataFrame = None,
    comments_df: pd.DataFrame = None,
    output_dir: str = "output",
    filename_prefix: str = "social_analysis",
) -> str:
    """
    Build a social media analysis Excel workbook.

    Sheets:
    - Posts: all posts with engagement metrics
    - Sentiment: post-level sentiment aggregations
    - Accounts: platform account summaries
    - Comments: individual comment text, sentiment, and like count (if fetched)
    - Commentary: AI-generated summaries (if provided)
    """
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"{filename_prefix}_{ts}.xlsx")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        write_sheet(writer, posts_df, "Posts")

        if sentiment_df is not None and not sentiment_df.empty:
            write_sheet(writer, sentiment_df, "Sentiment")

        if accounts_df is not None and not accounts_df.empty:
            write_sheet(writer, accounts_df, "Accounts")

        if comments_df is not None and not comments_df.empty:
            write_sheet(writer, comments_df, "Comments", wrap_columns=["comment_text"])

        if commentary_df is not None and not commentary_df.empty:
            write_sheet(writer, commentary_df, "Commentary", wrap_columns=["commentary"])

    logger.info("Social workbook saved: %s", path)
    return path
